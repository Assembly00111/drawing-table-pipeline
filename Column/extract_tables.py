"""
extract_tables.py
-----------------
Batch-extracts tabular data from engineering drawing screenshots using
Llama 3.2 Vision via Ollama (100% local, completely free).

SETUP
-----
1. Install Ollama:
       https://ollama.com/download
       (available for Windows, Mac, Linux)

2. Pull the vision model (one-time, ~6 GB download):
       ollama pull llama3.2-vision

3. Install Python dependencies:
       pip install ollama openpyxl pillow

4. Make sure Ollama is running before executing this script:
       ollama serve        <- run this in a separate terminal if needed
       (on Windows/Mac it usually runs automatically in the system tray)

USAGE
-----
    python extract_tables.py --input ./screenshots --output material_list.xlsx

Arguments:
  --input   Folder with PNG/JPG/JPEG screenshots (default: ./screenshots)
  --output  Output Excel filename (default: material_list.xlsx)
  --model   Ollama vision model to use (default: llama3.2-vision)
"""

import base64
import json
import sys
import argparse
from pathlib import Path

import ollama
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Prompt ────────────────────────────────────────────────────────────────────
PROMPT = """Extract the material list table from this engineering drawing screenshot.
Return ONLY a JSON array of row objects with these exact keys:
mark, profile, material, no, length, area, weight

Rules:
- Skip header rows and total rows — data rows only
- "no" and "length" are integers; "area" and "weight" are floats
- If a cell is empty or unreadable, use null
- Return ONLY valid JSON, no markdown fences, no explanation

Example output:
[{"mark":"BC406","profile":"RHS300*200*10","material":"E350","no":1,"length":4676,"area":4.7,"weight":352.4}]"""


# ── Ollama call ───────────────────────────────────────────────────────────────
def extract_rows_from_image(image_path: Path, model: str) -> list[dict]:
    """Send one image to Ollama and return parsed row dicts."""
    with open(image_path, "rb") as f:
        image_data = base64.b64encode(f.read()).decode("utf-8")

    response = ollama.chat(
        model=model,
        messages=[
            {
                "role": "user",
                "content": PROMPT,
                "images": [image_data],
            }
        ],
    )

    raw = response["message"]["content"].strip()
    # Strip markdown fences if model adds them anyway
    if raw.startswith("```"):
        raw = raw.split("```")[1]
        if raw.startswith("json"):
            raw = raw[4:]
    return json.loads(raw.strip())


# ── Excel writer ──────────────────────────────────────────────────────────────
def write_excel(all_rows: list[dict], output_path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Material List"

    thin    = Side(style="thin", color="CCCCCC")
    border  = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_font  = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    hdr_fill  = PatternFill("solid", start_color="404040")
    data_font = Font(name="Arial", size=10)
    src_font  = Font(name="Arial", size=9, italic=True, color="808080")
    tot_font  = Font(name="Arial", bold=True, size=10)
    tot_fill  = PatternFill("solid", start_color="D9D9D9")

    def bc(cell):
        cell.border = border

    # Header
    headers = ["Mark", "Profile", "Material", "No.", "Length", "Area", "Weight", "Source File"]
    for col, label in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=label)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        bc(c)
    ws.row_dimensions[1].height = 20

    # Data rows
    for i, row in enumerate(all_rows, 2):
        fill = PatternFill("solid", start_color="FFFFFF" if i % 2 == 0 else "F2F2F2")
        values = [
            row.get("mark"), row.get("profile"), row.get("material"),
            row.get("no"), row.get("length"), row.get("area"),
            row.get("weight"), row.get("_source", ""),
        ]
        for col, val in enumerate(values, 1):
            c = ws.cell(row=i, column=col, value=val)
            c.font = src_font if col == 8 else data_font
            c.fill = fill
            c.alignment = Alignment(
                horizontal="left" if col in (1, 2, 3, 8) else "center",
                vertical="center",
            )
            bc(c)
        ws.row_dimensions[i].height = 18

    # Total row
    tr = len(all_rows) + 2
    ws.merge_cells(f"A{tr}:E{tr}")
    c = ws.cell(row=tr, column=1, value="TOTAL")
    c.font = tot_font; c.fill = tot_fill
    c.alignment = Alignment(horizontal="right", vertical="center"); bc(c)
    for col, letter in [(6, "F"), (7, "G")]:
        c = ws.cell(row=tr, column=col, value=f"=SUM({letter}2:{letter}{tr-1})")
        c.font = tot_font; c.fill = tot_fill; c.number_format = "0.0"
        c.alignment = Alignment(horizontal="center", vertical="center"); bc(c)

    # Column widths
    for i, w in enumerate([10, 18, 12, 7, 10, 8, 10, 22], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(output_path)
    print(f"\n✅  Saved → {output_path}  ({len(all_rows)} data rows)")


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="Batch screenshot tables → Excel (via Ollama)")
    parser.add_argument("--input",  default="./screenshots", help="Folder with image files")
    parser.add_argument("--output", default="material_list.xlsx", help="Output Excel file")
    parser.add_argument("--model",  default="llama3.2-vision", help="Ollama model name")
    args = parser.parse_args()

    # Check Ollama is reachable
    try:
        ollama.list()
    except Exception:
        sys.exit(
            "❌  Cannot reach Ollama. Make sure it is installed and running.\n"
            "    Download: https://ollama.com/download\n"
            "    Then run: ollama pull llama3.2-vision"
        )

    input_dir = Path(args.input)
    if not input_dir.is_dir():
        sys.exit(f"❌  Input folder not found: {input_dir}")

    images = sorted(
        p for p in input_dir.iterdir()
        if p.suffix.lower() in (".png", ".jpg", ".jpeg")
    )
    if not images:
        sys.exit(f"❌  No PNG/JPG images found in {input_dir}")

    print(f"Model  : {args.model}")
    print(f"Images : {len(images)} file(s) in '{input_dir}'")
    print(f"Output : {args.output}\n")

    all_rows, errors = [], []

    for idx, img_path in enumerate(images, 1):
        print(f"[{idx}/{len(images)}] {img_path.name} ... ", end="", flush=True)
        try:
            rows = extract_rows_from_image(img_path, args.model)
            for r in rows:
                r["_source"] = img_path.name
            all_rows.extend(rows)
            print(f"✓  {len(rows)} row(s)")
        except Exception as e:
            print(f"✗  ERROR: {e}")
            errors.append((img_path.name, str(e)))

    if not all_rows:
        sys.exit("❌  No data extracted. Check your images and that the model is pulled.")

    write_excel(all_rows, args.output)

    if errors:
        print(f"\n⚠️  {len(errors)} file(s) failed:")
        for name, err in errors:
            print(f"   • {name}: {err}")


if __name__ == "__main__":
    main()
