"""
extract_tables.py
-----------------
Batch-extracts tabular data from engineering drawing screenshots using
a vision model via Ollama — optimized for low VRAM (6GB GPUs).

Output Excel:
  Sheet 1 "Material List" — all rows grouped by assembly, with merged
           Assembly/Qty columns, plus Total Area and Total Weight columns.
  Sheet 2 "Summary"       — one row per assembly: assembly name, qty,
           profile, total area, total weight. Grand total at the bottom.

SETUP
-----
1. Install Ollama: https://ollama.com/download
2. Pull a vision model:
       ollama pull granite3.2-vision      # recommended: 2B, fits in 6GB
       ollama pull qwen2.5vl:3b           # alternative
3. Install Python dependencies:
       pip install ollama openpyxl pillow

USAGE
-----
    python extract_tables.py --input ./screenshots --output material_list.xlsx --model granite3.2-vision

Arguments:
  --input    Folder with PNG/JPG/JPEG screenshots (default: ./screenshots)
  --output   Output Excel filename (default: material_list.xlsx)
  --model    Ollama vision model (default: granite3.2-vision)
  --start    Start index, 1-based (default: 1)
  --end      End index, 1-based inclusive (default: all)
"""

import base64
import json
import sys
import argparse
from pathlib import Path
from io import BytesIO

import ollama
from PIL import Image
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Prompt ────────────────────────────────────────────────────────────────────
PROMPT = """Extract the material list table from this engineering drawing screenshot.
Also extract the assembly name (shown next to 'MATERIAL LIST FOR ASSEMBLY MK\'D') and
the quantity required (the number shown after the assembly name, before 'No. Required').

Return ONLY a JSON object with this exact structure:
{
  "assembly": "BBR1",
  "qty": 1,
  "rows": [
    {"mark":"BBR1","profile":"SHS132*8","material":"E350","no":1,"length":892,"area":0.4,"weight":25.5}
  ]
}

Rules:
- "assembly" is the assembly mark (e.g. TR201, TBR3, BBR1)
- "qty" is the integer quantity required (No. Required field)
- "rows" contains only data rows — skip header rows and total rows
- In rows: "no" and "length" are integers; "area" and "weight" are floats
- If a cell is empty or unreadable, use null
- Return ONLY valid JSON, no markdown fences, no explanation"""


# ── Image preprocessing ───────────────────────────────────────────────────────
def preprocess_image(image_path: Path) -> str:
    from PIL import ImageOps
    img = Image.open(image_path).convert("L")   # grayscale
    img = ImageOps.invert(img)                   # white-on-dark → black-on-white
    buf = BytesIO()
    img.save(buf, format="PNG", optimize=True)
    return base64.b64encode(buf.getvalue()).decode("utf-8")


# ── Ollama call ───────────────────────────────────────────────────────────────
def extract_from_image(image_path: Path, model: str) -> dict:
    image_data = preprocess_image(image_path)
    response = ollama.chat(
        model=model,
        messages=[{"role": "user", "content": PROMPT, "images": [image_data]}],
        options={"temperature": 0, "num_predict": 768, "num_ctx": 2048},
    )
    raw = response["message"]["content"].strip()
    if raw.startswith("```"):
        raw = raw.split("```")[1]
        if raw.startswith("json"):
            raw = raw[4:]
    return json.loads(raw.strip())


# ── Shared styles ─────────────────────────────────────────────────────────────
def make_styles():
    thin    = Side(style="thin",   color="CCCCCC")
    med     = Side(style="medium", color="999999")
    return {
        "thin": thin, "med": med,
        "brd":     Border(left=thin, right=thin, top=thin, bottom=thin),
        "grp_brd": Border(left=med,  right=med,  top=med,  bottom=med),
        "hdr_font":  Font(name="Arial", bold=True, size=10, color="FFFFFF"),
        "hdr_fill":  PatternFill("solid", start_color="1F3864"),
        "ext_hdr_fill": PatternFill("solid", start_color="1B4F72"),
        "asm_font":  Font(name="Arial", bold=True, size=10, color="FFFFFF"),
        "asm_fill":  PatternFill("solid", start_color="2E4057"),
        "data_font": Font(name="Arial", size=10),
        "src_font":  Font(name="Arial", size=9, italic=True, color="888888"),
        "ext_font":  Font(name="Arial", size=10, color="1A5276"),
        "wht_fill":  PatternFill("solid", start_color="FFFFFF"),
        "alt_fill":  PatternFill("solid", start_color="F2F2F2"),
        "ext_e":     PatternFill("solid", start_color="EBF5FB"),
        "ext_o":     PatternFill("solid", start_color="D6EAF8"),
        "tot_font":  Font(name="Arial", bold=True, size=10),
        "tot_fill":  PatternFill("solid", start_color="D9D9D9"),
        "ext_tot_fill": PatternFill("solid", start_color="AED6F1"),
        "ext_tot_font": Font(name="Arial", bold=True, size=10, color="1A5276"),
        "ctr": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "lft": Alignment(horizontal="left",   vertical="center"),
        "rgt": Alignment(horizontal="right",  vertical="center"),
    }


# ── Sheet 1: Material List ────────────────────────────────────────────────────
def write_material_sheet(ws, groups: list[dict], s: dict):
    def bc(cell, b=None): cell.border = b or s["brd"]

    # Columns: A=Assembly B=Qty C=Mark D=Profile E=Material F=No. G=Length
    #          H=Area I=Weight J=Source K=Total Area L=Total Weight
    headers = ["Assembly\n(MK'D)", "Qty\nReq'd", "Mark", "Profile", "Material",
               "No.", "Length", "Area", "Weight", "Source File",
               "Total Area\n(No.×Area)", "Total Weight\n(No.×Weight)"]
    col_w   = [14, 7, 10, 18, 12, 7, 10, 8, 10, 26, 13, 14]

    for col, label in enumerate(headers, 1):
        c = ws.cell(1, col, value=label)
        c.fill = s["ext_hdr_fill"] if col >= 11 else s["hdr_fill"]
        c.font = s["hdr_font"]; c.alignment = s["ctr"]; bc(c)
    ws.row_dimensions[1].height = 30
    for i, w in enumerate(col_w, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    current_row = 2
    for g_idx, grp in enumerate(groups):
        rows     = grp.get("rows", [])
        asm_name = grp.get("assembly") or "?"
        qty      = grp.get("qty") or 1
        src      = grp.get("_source", "")
        n        = len(rows)
        if n == 0:
            continue

        start_row = current_row
        end_row   = current_row + n - 1
        fill      = s["wht_fill"] if g_idx % 2 == 0 else s["alt_fill"]
        efill     = s["ext_e"]    if g_idx % 2 == 0 else s["ext_o"]

        for i, r in enumerate(rows):
            rn = current_row + i
            ws.cell(rn, 1).fill = s["asm_fill"]
            ws.cell(rn, 2).fill = s["asm_fill"]

            vals = [r.get("mark"), r.get("profile"), r.get("material"),
                    r.get("no"), r.get("length"), r.get("area"), r.get("weight"), src]
            for col, val in enumerate(vals, 3):
                c = ws.cell(rn, col, value=val)
                c.font      = s["src_font"] if col == 10 else s["data_font"]
                c.fill      = fill
                c.alignment = s["lft"] if col in (3,4,5,10) else s["ctr"]
                bc(c)

            # K: Total Area  = No. * Area
            ck = ws.cell(rn, 11, value=f"=F{rn}*H{rn}")
            ck.font = s["ext_font"]; ck.fill = efill
            ck.number_format = "0.00"; ck.alignment = s["ctr"]; bc(ck)

            # L: Total Weight = No. * Weight
            cl = ws.cell(rn, 12, value=f"=F{rn}*I{rn}")
            cl.font = s["ext_font"]; cl.fill = efill
            cl.number_format = "0.0"; cl.alignment = s["ctr"]; bc(cl)

            ws.row_dimensions[rn].height = 18

        if n > 1:
            ws.merge_cells(f"A{start_row}:A{end_row}")
            ws.merge_cells(f"B{start_row}:B{end_row}")

        a = ws.cell(start_row, 1, value=asm_name)
        a.font = s["asm_font"]; a.fill = s["asm_fill"]; a.alignment = s["ctr"]; bc(a, s["grp_brd"])
        q = ws.cell(start_row, 2, value=qty)
        q.font = s["asm_font"]; q.fill = s["asm_fill"]; q.alignment = s["ctr"]; bc(q, s["grp_brd"])

        for rn in range(start_row, end_row + 1):
            for col in range(3, 13):
                ws.cell(rn, col).border = Border(
                    left   = s["med"] if col == 3  else s["thin"],
                    right  = s["med"] if col == 12 else s["thin"],
                    top    = s["med"] if rn == start_row else s["thin"],
                    bottom = s["med"] if rn == end_row   else s["thin"],
                )
        current_row += n

    # Total row
    tr = current_row
    ws.merge_cells(f"A{tr}:G{tr}")
    c = ws.cell(tr, 1, value="TOTAL")
    c.font = s["tot_font"]; c.fill = s["tot_fill"]; c.alignment = s["rgt"]; bc(c)
    for col, letter, fmt in [(8,"H","0.00"),(9,"I","0.0"),(11,"K","0.00"),(12,"L","0.0")]:
        c = ws.cell(tr, col, value=f"=SUM({letter}2:{letter}{tr-1})")
        c.font = s["ext_tot_font"] if col >= 11 else s["tot_font"]
        c.fill = s["ext_tot_fill"] if col >= 11 else s["tot_fill"]
        c.number_format = fmt; c.alignment = s["ctr"]; bc(c)
    c = ws.cell(tr, 10); c.fill = s["tot_fill"]; bc(c)
    ws.row_dimensions[tr].height = 20
    ws.freeze_panes = "A2"


# ── Sheet 2: Summary ──────────────────────────────────────────────────────────
def write_summary_sheet(ws2, groups: list[dict], s: dict):
    def sbc(cell, b=None): cell.border = b or s["brd"]

    grand_fill = PatternFill("solid", start_color="1F3864")
    grand_font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    sub_fill   = PatternFill("solid", start_color="2E4057")

    # Title
    ws2.merge_cells("A1:E1")
    t = ws2.cell(1, 1, value="MEMBER SUMMARY — TOTAL AREA & WEIGHT")
    t.font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    t.fill = PatternFill("solid", start_color="1F3864")
    t.alignment = s["ctr"]; sbc(t)
    ws2.row_dimensions[1].height = 26

    # Sub-headers
    sub_hdrs = ["Assembly\n(MK'D)", "Qty\nReq'd", "Profile", "Total Area (m²)", "Total Weight (kg)"]
    sub_w    = [16, 9, 18, 18, 18]
    for col, (lbl, w) in enumerate(zip(sub_hdrs, sub_w), 1):
        c = ws2.cell(2, col, value=lbl)
        c.font = s["hdr_font"]; c.fill = sub_fill; c.alignment = s["ctr"]; sbc(c)
        ws2.column_dimensions[get_column_letter(col)].width = w
    ws2.row_dimensions[2].height = 28

    # Data rows — one per assembly
    s_row = 3
    for g_idx, grp in enumerate(groups):
        rows     = grp.get("rows", [])
        asm_name = grp.get("assembly") or "?"
        qty      = grp.get("qty") or 1
        profile  = rows[0].get("profile") if rows else ""
        total_area   = sum((r.get("no") or 0) * (r.get("area")   or 0) for r in rows)
        total_weight = sum((r.get("no") or 0) * (r.get("weight") or 0) for r in rows)

        fill = s["wht_fill"] if g_idx % 2 == 0 else s["alt_fill"]
        for col, val in enumerate([asm_name, qty, profile,
                                    round(total_area, 2), round(total_weight, 1)], 1):
            c = ws2.cell(s_row, col, value=val)
            c.font = s["data_font"]; c.fill = fill
            c.alignment = s["lft"] if col == 3 else s["ctr"]; sbc(c)
        ws2.row_dimensions[s_row].height = 18
        s_row += 1

    # Grand total
    ws2.merge_cells(f"A{s_row}:C{s_row}")
    c = ws2.cell(s_row, 1, value="GRAND TOTAL")
    c.font = grand_font; c.fill = grand_fill; c.alignment = s["rgt"]; sbc(c)
    for col, letter, fmt in [(4,"D","0.00"),(5,"E","0.0")]:
        c = ws2.cell(s_row, col, value=f"=SUM({letter}3:{letter}{s_row-1})")
        c.font = grand_font; c.fill = grand_fill
        c.number_format = fmt; c.alignment = s["ctr"]; sbc(c)
    ws2.row_dimensions[s_row].height = 22
    ws2.freeze_panes = "A3"


# ── Write workbook ────────────────────────────────────────────────────────────
def write_excel(groups: list[dict], output_path: str):
    wb  = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Material List"
    ws2 = wb.create_sheet("Summary")

    s = make_styles()
    write_material_sheet(ws1, groups, s)
    write_summary_sheet(ws2, groups, s)

    # Make Summary the active sheet when file opens
    wb.active = ws2

    wb.save(output_path)
    total_rows = sum(len(g.get("rows", [])) for g in groups)
    print(f"\n✅  Saved → {output_path}")
    print(f"    Material List : {total_rows} data rows, {len(groups)} assemblies")
    print(f"    Summary       : {len(groups)} assembly rows + grand total")


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="Batch screenshot tables → Excel with summary")
    parser.add_argument("--input",  default="./screenshots")
    parser.add_argument("--output", default="material_list.xlsx")
    parser.add_argument("--model",  default="granite3.2-vision")
    parser.add_argument("--start",  type=int, default=1)
    parser.add_argument("--end",    type=int, default=None)
    args = parser.parse_args()

    try:
        ollama.list()
    except Exception:
        sys.exit("❌  Cannot reach Ollama. Make sure it is installed and running.")

    input_dir = Path(args.input)
    if not input_dir.is_dir():
        sys.exit(f"❌  Input folder not found: {input_dir}")

    all_images = sorted(p for p in input_dir.iterdir()
                        if p.suffix.lower() in (".png", ".jpg", ".jpeg"))
    if not all_images:
        sys.exit(f"❌  No images found in {input_dir}")

    start_idx = max(1, args.start) - 1
    end_idx   = args.end if args.end else len(all_images)
    images    = all_images[start_idx:end_idx]

    print(f"Model   : {args.model}")
    print(f"Images  : {len(images)} file(s)  (#{start_idx+1}–#{start_idx+len(images)} of {len(all_images)})")

    groups, errors = [], []

    for idx, img_path in enumerate(images, 1):
        print(f"[{idx}/{len(images)}] {img_path.name} ... ", end="", flush=True)
        try:
            result = extract_from_image(img_path, args.model)
            result["_source"] = img_path.name
            groups.append(result)
            print(f"✓  assembly={result.get('assembly')}  qty={result.get('qty')}  {len(result.get('rows',[]))} row(s)")
        except Exception as e:
            print(f"✗  ERROR: {e}")
            errors.append((img_path.name, str(e)))

    if not groups:
        sys.exit("❌  No data extracted.")

    write_excel(groups, args.output)

    if errors:
        print(f"\n⚠️  {len(errors)} file(s) failed:")
        for name, err in errors:
            print(f"   • {name}: {err}")

if __name__ == "__main__":
    main()
