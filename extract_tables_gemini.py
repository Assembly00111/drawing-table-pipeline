"""
extract_tables_gemini.py
------------------------
Batch-extracts tabular data from engineering drawing screenshots using
Google Gemini Vision API (free tier: 1,500 requests/day).
Each screenshot's data is saved in a separate sheet in the same workbook.

SETUP
-----
1. Get a free API key (no credit card needed):
       https://aistudio.google.com
       Click "Get API Key" → "Create API key"

2. Install Python dependencies:
       pip install google-genai openpyxl pillow

3. Set your API key:
       Windows :  set GEMINI_API_KEY=AIza...
       Mac/Linux: export GEMINI_API_KEY=AIza...

USAGE
-----
    python extract_tables_gemini.py --input ./screenshots --output material_list.xlsx

Arguments:
  --input   Folder with PNG/JPG/JPEG screenshots (default: ./screenshots)
  --output  Output Excel filename (default: material_list.xlsx)
  --delay   Seconds between API calls (default: 1.0 — stay within free rate limits)
"""

import base64
import json
import os
import sys
import time
import argparse
from pathlib import Path

from google import genai
from google.genai import types
from PIL import Image
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Prompt ────────────────────────────────────────────────────────────────────
PROMPT = """Extract the material list table from this engineering drawing screenshot.
Return ONLY a JSON array of row objects with these exact keys:
mark, profile, material, no, length, area, weight

Rules:
- Skip header rows and total rows — data rows only
- "no" and "length" are integers; "area" and "weight" are floats
- If a cell is empty or unreadable, use null
- Return ONLY valid JSON, no markdown fences, no explanation

Example output:
[{"mark":"BC406","profile":"RHS300*200*10","material":"E350","no":1,"length":4676,"area":4.7,"weight":352.4}]"""


# ── Gemini call ───────────────────────────────────────────────────────────────
def extract_rows_from_image(image_path: Path, client: genai.Client) -> list[dict]:
    """Send one image to Gemini and return parsed row dicts."""
    with open(image_path, "rb") as f:
        image_bytes = f.read()

    suffix = image_path.suffix.lower()
    mime_map = {".png": "image/png", ".jpg": "image/jpeg", ".jpeg": "image/jpeg"}
    mime_type = mime_map.get(suffix, "image/png")

    response = client.models.generate_content(
        model="gemini-3.5-flash",
        contents=[
            types.Part.from_bytes(data=image_bytes, mime_type=mime_type),
            PROMPT,
        ],
    )

    raw = response.text.strip()
    # Strip markdown fences if model adds them
    if raw.startswith("```"):
        raw = raw.split("```")[1]
        if raw.startswith("json"):
            raw = raw[4:]
    return json.loads(raw.strip())


# ── Sheet writer ──────────────────────────────────────────────────────────────
def make_sheet_name(filename: str, index: int, existing: set) -> str:
    """Create a unique, Excel-safe sheet name from the filename."""
    # Remove extension, truncate to 28 chars (leaving room for dedup suffix)
    name = Path(filename).stem[:28]
    # Excel forbids these characters in sheet names
    for ch in r'\/?*[]':
        name = name.replace(ch, "_")
    # Ensure uniqueness
    candidate = name
    counter = 2
    while candidate in existing:
        candidate = f"{name[:25]}_{counter}"
        counter += 1
    existing.add(candidate)
    return candidate


def write_sheet(ws, rows: list[dict], source_name: str):
    """Write rows for one screenshot into a worksheet."""
    thin      = Side(style="thin", color="CCCCCC")
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_font  = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    hdr_fill  = PatternFill("solid", start_color="404040")
    data_font = Font(name="Arial", size=10)
    tot_font  = Font(name="Arial", bold=True, size=10)
    tot_fill  = PatternFill("solid", start_color="D9D9D9")

    def bc(cell):
        cell.border = border

    # Source file label
    ws.cell(row=1, column=1, value=f"Source: {source_name}").font = Font(
        name="Arial", size=9, italic=True, color="808080"
    )
    ws.row_dimensions[1].height = 16

    # Header row
    headers = ["Mark", "Profile", "Material", "No.", "Length", "Area", "Weight"]
    for col, label in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=label)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        bc(c)
    ws.row_dimensions[2].height = 20

    # Data rows
    for i, row in enumerate(rows, 3):
        fill = PatternFill("solid", start_color="FFFFFF" if i % 2 == 0 else "F2F2F2")
        values = [
            row.get("mark"), row.get("profile"), row.get("material"),
            row.get("no"), row.get("length"), row.get("area"), row.get("weight"),
        ]
        for col, val in enumerate(values, 1):
            c = ws.cell(row=i, column=col, value=val)
            c.font = data_font
            c.fill = fill
            c.alignment = Alignment(
                horizontal="left" if col in (1, 2, 3) else "center",
                vertical="center",
            )
            bc(c)
        ws.row_dimensions[i].height = 18

    # Total row
    tr = len(rows) + 3
    ws.merge_cells(f"A{tr}:E{tr}")
    c = ws.cell(row=tr, column=1, value="TOTAL")
    c.font = tot_font; c.fill = tot_fill
    c.alignment = Alignment(horizontal="right", vertical="center"); bc(c)
    for col, letter in [(6, "F"), (7, "G")]:
        c = ws.cell(row=tr, column=col, value=f"=SUM({letter}3:{letter}{tr-1})")
        c.font = tot_font; c.fill = tot_fill; c.number_format = "0.0"
        c.alignment = Alignment(horizontal="center", vertical="center"); bc(c)

    # Column widths
    for i, w in enumerate([10, 18, 12, 7, 10, 8, 10], 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_summary_sheet(wb, all_rows: list[dict]):
    """Write a summary sheet with all rows combined."""
    ws = wb.create_sheet("SUMMARY", 0)  # Insert at front

    thin      = Side(style="thin", color="CCCCCC")
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_font  = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    hdr_fill  = PatternFill("solid", start_color="1F3864")
    data_font = Font(name="Arial", size=10)
    src_font  = Font(name="Arial", size=9, italic=True, color="808080")
    tot_font  = Font(name="Arial", bold=True, size=10)
    tot_fill  = PatternFill("solid", start_color="D9D9D9")

    def bc(cell):
        cell.border = border

    headers = ["Mark", "Profile", "Material", "No.", "Length", "Area", "Weight", "Source File"]
    for col, label in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=label)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        bc(c)
    ws.row_dimensions[1].height = 20

    for i, row in enumerate(all_rows, 2):
        fill = PatternFill("solid", start_color="FFFFFF" if i % 2 == 0 else "F2F2F2")
        values = [
            row.get("mark"), row.get("profile"), row.get("material"),
            row.get("no"), row.get("length"), row.get("area"),
            row.get("weight"), row.get("_source", ""),
        ]
        for col, val in enumerate(values, 1):
            c = ws.cell(row=i, column=col, value=val)
            c.font = src_font if col == 8 else data_font
            c.fill = fill
            c.alignment = Alignment(
                horizontal="left" if col in (1, 2, 3, 8) else "center",
                vertical="center",
            )
            bc(c)
        ws.row_dimensions[i].height = 18

    tr = len(all_rows) + 2
    ws.merge_cells(f"A{tr}:E{tr}")
    c = ws.cell(row=tr, column=1, value="TOTAL")
    c.font = tot_font; c.fill = tot_fill
    c.alignment = Alignment(horizontal="right", vertical="center"); bc(c)
    for col, letter in [(6, "F"), (7, "G")]:
        c = ws.cell(row=tr, column=col, value=f"=SUM({letter}2:{letter}{tr-1})")
        c.font = tot_font; c.fill = tot_fill; c.number_format = "0.0"
        c.alignment = Alignment(horizontal="center", vertical="center"); bc(c)

    for i, w in enumerate([10, 18, 12, 7, 10, 8, 10, 28], 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="Batch screenshot tables → Excel (via Gemini)")
    parser.add_argument("--input",  default="./screenshots", help="Folder with image files")
    parser.add_argument("--output", default="material_list.xlsx", help="Output Excel file")
    parser.add_argument("--delay",  type=float, default=1.0, help="Seconds between API calls")
    args = parser.parse_args()

    api_key = os.environ.get("GEMINI_API_KEY")
    if not api_key:
        sys.exit(
            "❌  GEMINI_API_KEY environment variable not set.\n"
            "    Get a free key at: https://aistudio.google.com\n"
            "    Then run:  set GEMINI_API_KEY=AIza...    (Windows)\n"
            "               export GEMINI_API_KEY=AIza... (Mac/Linux)"
        )

    input_dir = Path(args.input)
    if not input_dir.is_dir():
        sys.exit(f"❌  Input folder not found: {input_dir}")

    images = sorted(
        p for p in input_dir.iterdir()
        if p.suffix.lower() in (".png", ".jpg", ".jpeg")
    )
    if not images:
        sys.exit(f"❌  No PNG/JPG images found in {input_dir}")

    client = genai.Client(api_key=api_key)

    print(f"Model  : gemini-3.5-flash (free tier)")
    print(f"Images : {len(images)} file(s) in '{input_dir}'")
    print(f"Output : {args.output}\n")

    wb         = openpyxl.Workbook()
    wb.remove(wb.active)   # remove default empty sheet
    all_rows   = []
    errors     = []
    used_names = set()

    for idx, img_path in enumerate(images, 1):
        print(f"[{idx}/{len(images)}] {img_path.name} ... ", end="", flush=True)
        try:
            rows = extract_rows_from_image(img_path, client)
            for r in rows:
                r["_source"] = img_path.name
            all_rows.extend(rows)

            sheet_name = make_sheet_name(img_path.name, idx, used_names)
            ws = wb.create_sheet(title=sheet_name)
            write_sheet(ws, rows, img_path.name)
            print(f"✓  {len(rows)} row(s)  → sheet '{sheet_name}'")
        except Exception as e:
            print(f"✗  ERROR: {e}")
            errors.append((img_path.name, str(e)))

        if idx < len(images):
            time.sleep(args.delay)

    if not all_rows:
        sys.exit("❌  No data extracted. Check your images and API key.")

    # Add summary sheet at the front
    write_summary_sheet(wb, all_rows)
    wb.save(args.output)
    print(f"\n✅  Saved → {args.output}")
    print(f"    Sheets: 1 SUMMARY + {len(images) - len(errors)} individual sheet(s)")
    print(f"    Total rows: {len(all_rows)}")

    if errors:
        print(f"\n⚠️  {len(errors)} file(s) failed:")
        for name, err in errors:
            print(f"   • {name}: {err}")


if __name__ == "__main__":
    main()
